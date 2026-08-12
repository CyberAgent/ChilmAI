from __future__ import annotations

import base64
import io
import json
import os
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Callable, Literal
from urllib.parse import quote, unquote, urlparse

import openpyxl

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, ConfigDict, Field

from apps.version import CHILMAI_VERSION
from chilmai.algorithm.cp_use_transfer.CP_algo import OptimizationFailureError
from chilmai.generic.column_mapper import detect_section
from chilmai.generic.config import (
    DEFAULT_CONFIG,
    DEFAULT_PROFILE_NAME,
    SCHEMA_VERSION,
    ConfigStore,
)
from chilmai.generic.error_codes import ChilmError, ErrorCode
from chilmai.generic.parser import InputParser
from chilmai.generic.service import MatchingService

APP_DIR = Path(__file__).resolve().parent
ASSETS_DIR = APP_DIR / "assets"

OPENAPI_TAGS = [
    {
        "name": "API",
        "description": "評価や連携検証向けの API です。",
    },
]

app = FastAPI(
    title="ChilmAI Generic API",
    summary="API for daycare matching logic",
    description=(
        "ChilmAI のサーバ API は、申込者データと保育所データを検証し、"
        "OR-Tools CP-SAT による保育所利用調整のマッチングを実行します。"
        "本番運用に必要な認証、認可、監査ログ、監視、組織ごとのセキュリティ要件は含みません。"
    ),
    version=CHILMAI_VERSION,
    openapi_tags=OPENAPI_TAGS,
    redoc_url=None,
)
app.mount("/static", StaticFiles(directory=APP_DIR / "static"), name="static")
app.mount("/assets", StaticFiles(directory=ASSETS_DIR), name="assets")
templates = Jinja2Templates(directory=str(APP_DIR / "templates"))
templates.env.globals["is_packaged"] = os.getenv("CHILMAI_PACKAGED") == "1"


def _profile_display(name: str | None) -> str:
    if not isinstance(name, str):
        return ""
    return "デフォルト" if name == DEFAULT_PROFILE_NAME else name


templates.env.filters["profile_display"] = _profile_display

config_store = ConfigStore("data/config.json")
matching_service = MatchingService()


def _infer_file_format_from_filename(filename: str | None) -> str:
    if not filename:
        raise ChilmError("Unsupported file format: ", code=ErrorCode.UNSUPPORTED_FILE_FORMAT)
    lower = filename.lower()
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith((".xlsx", ".xls", ".xlsm", ".xlsb")):
        return "excel"
    raise ChilmError(f"Unsupported file format: {filename}", code=ErrorCode.UNSUPPORTED_FILE_FORMAT)


def _get_config_display() -> dict[str, Any]:
    snapshot = config_store.snapshot()
    loaded = snapshot["active_config"]
    return {
        "is_customized": loaded != DEFAULT_CONFIG,
        "active_profile_name": snapshot["active_name"],
        "profile_names": snapshot["profile_names"],
    }


def _static_asset_version() -> str:
    candidates = [
        APP_DIR / "static" / "style.css",
        APP_DIR / "static" / "htmx-lite.js",
    ]
    latest = max(int(path.stat().st_mtime) for path in candidates if path.exists())
    return str(latest)


def _display_value(value: Any) -> str:
    if value is None:
        return "未割当"
    if isinstance(value, bool):
        return "True" if value else "False"
    if isinstance(value, (list, tuple)):
        if not value:
            return "-"
        return ", ".join(_display_value(item) for item in value)
    if isinstance(value, dict):
        if not value:
            return "-"
        return ", ".join(f"{key}: {_display_value(item)}" for key, item in value.items())
    return str(value)


_SUMMARY_LABELS: dict[str, str] = {
    "children_count": "申込者数",
    "daycares_count": "保育所数",
}

_CHILDREN_LABELS: dict[str, str] = {
    "child_id": "申請者ID",
    "household_id": "世帯ID",
    "age": "年齢区分",
    "score": "点数列名（全施設共通の場合）",
    "score_prefix": "点数列名のパターン（施設ごとに点数が異なる場合）",
    "preference_prefix": "希望保育園ID列のパターン",
    "enrolled_daycare_id": "在籍保育所ID",
    "sibling_pattern": "兄弟姉妹パターン",
}

_CHILDREN_SAMPLES: dict[str, str] = {
    "child_id": "例: 申請者番号",
    "household_id": "例: 世帯番号",
    "age": "例: 年齢",
    "score": "例: 点数（全施設共通。「点数列名のパターン（施設ごとに点数が異なる場合）」との併用不可）",
    "score_prefix": "N が数字の位置を表します。先頭: 「N希望点数」→ 1希望点数, 2希望点数 ... / 途中: 「第N希望点数」→ 第1希望点数, 第2希望点数 ... / 末尾: 「点数N」→ 点数1, 点数2 ...（数字は半角）",
    "preference_prefix": "N が数字の位置を表します。先頭: 「N番目希望」→ 1番目希望, 2番目希望 ... / 途中: 「第N希望保育園」→ 第1希望保育園, 第2希望保育園 ... / 末尾: 「希望保育園ID_N」→ 希望保育園ID_1, 希望保育園ID_2 ...（数字は半角）",
    "enrolled_daycare_id": "例: 在籍保育所ID",
    "sibling_pattern": "例: きょうだいパターン",
}

_DAYCARES_LABELS: dict[str, str] = {
    "daycare_id": "保育所ID",
    "daycare_name": "保育所名",
    "capacity_prefix": "募集人数列のパターン",
}

_DAYCARES_SAMPLES: dict[str, str] = {
    "daycare_id": "例: 保育所ID",
    "daycare_name": "例: 保育所名",
    "capacity_prefix": "N が数字の位置を表します。先頭: 「N歳募集人数」→ 0歳募集人数, 1歳募集人数 ... / 途中: 「定員N歳枠」→ 定員0歳枠, 定員1歳枠 ... / 末尾: 「募集人数N」→ 募集人数0, 募集人数1 ...（数字は半角）",
}

_COMBINATION_LABELS: dict[str, str] = {
    "household_id": "世帯ID",
    "rank": "組み合わせ順位列名",
    "child_code_prefix": "宛名コード列のパターン",
    "facility_prefix": "希望保育所ID列のパターン",
}

_COMBINATION_SAMPLES: dict[str, str] = {
    "household_id": "例: ファミリーコード",
    "rank": "例: 総当たり順位",
    "child_code_prefix": "N が数字の位置を表します。例: 「宛名コードN」→ 宛名コード1, 宛名コード2, 宛名コード3 ...（数字は半角）",
    "facility_prefix": "N が数字の位置を表します。例: 「希望施設N」→ 希望施設1, 希望施設2, 希望施設3 ...（数字は半角）",
}

_OUTPUT_LABELS: dict[str, str] = {
    "result_daycare_id": "入所選考結果保育所ID",
    "result_daycare_name": "入所選考結果保育所名",
}

_OUTPUT_SAMPLES: dict[str, str] = {
    "result_daycare_id": "例: 入所選考結果保育所ID",
    "result_daycare_name": "例: 入所選考結果保育所名",
}


def _error_type_for_code(code: int | None) -> Literal["config", "data", "format"]:
    if code is not None and 100 <= code < 200:
        return "config"
    if code is not None and 400 <= code < 500:
        return "format"
    return "data"


def _format_error_message(raw: str, code: int | None = None) -> str:
    if code == ErrorCode.UNSUPPORTED_FILE_FORMAT:
        prefix = "Unsupported file format: "
        filename = raw[len(prefix) :] if raw.startswith(prefix) else raw
        return f"「{filename}」はアップロードに対応していない形式です"
    return raw


def _labeled_error(msg: str, code: int | None) -> str:
    return f"[E{code}] {msg}" if code else msg


def _build_validate_view(result: dict[str, Any]) -> dict[str, Any]:
    config_errors = [
        _labeled_error(e["message"], e.get("code"))
        for e in result.get("errors", [])
        if e.get("type") == "config"
    ]
    data_errors = [
        _labeled_error(e["message"], e.get("code"))
        for e in result.get("errors", [])
        if e.get("type") == "data"
    ]
    format_errors = [
        _labeled_error(_format_error_message(e["message"], e.get("code")), e.get("code"))
        for e in result.get("errors", [])
        if e.get("type") == "format"
    ]
    has_warnings = bool(result.get("warnings"))
    if result["is_valid"]:
        status = "問題ありません（注意事項あり）" if has_warnings else "問題ありません"
        next_action = "確認の結果、問題はありませんでした。このままマッチングを実行してください。"
    elif format_errors:
        status = "問題があります"
        next_action = "ファイル形式または列名の構造に問題があります。ファイル形式または列名を確認し、必要に応じて修正して再アップロードしてください。"
    else:
        status = "問題があります"
        next_action = "確認の結果、エラーがありました。内容を修正してファイルを再アップロードしてください。"
    return {
        "status": status,
        "next_action": next_action,
        "config_errors": config_errors,
        "data_errors": data_errors,
        "format_errors": format_errors,
        "summary_rows": [
            {"label": _SUMMARY_LABELS.get(str(k), str(k)), "value": _display_value(v)}
            for k, v in result.get("summary", {}).items()
        ],
    }


def _id_sort_key(s: str) -> tuple[int, int, str]:
    """数値IDは数値順、非数値IDは文字順でソートする。"""
    try:
        return (0, int(s), "")
    except ValueError:
        return (1, 0, s)


def _match_summary(result: dict[str, Any], total_children: int) -> str:
    matched = result.get("matched_children", {}).get("total", 0)
    percentage = round(matched / total_children * 100) if total_children > 0 else 0
    summary = f"{total_children}人中{matched}人（{percentage}%）に割り当てました。"
    transfer_back_count = result.get("transfer_back_count", 0)
    if transfer_back_count > 0:
        summary += f"（別途 {transfer_back_count}人が転園元保育所に戻りました）"
    return summary


def _percentage(numerator: int, denominator: int) -> int:
    return round(numerator / denominator * 100) if denominator > 0 else 0


def _build_breakdown(matched_children: dict[str, Any]) -> dict[str, Any]:
    by_age_raw = matched_children.get("by_age", {}) or {}
    age_rows: list[dict[str, int]] = []
    for age in range(6):
        bucket = by_age_raw.get(str(age), {})
        applied = int(bucket.get("applied", 0))
        matched = int(bucket.get("matched", 0))
        age_rows.append(
            {
                "age": age,
                "applied": applied,
                "matched": matched,
                "percentage": _percentage(matched, applied),
            }
        )

    total_applied = int(matched_children.get("applied_total", 0))
    total_matched = int(matched_children.get("total", 0))

    return {
        "age_rows": age_rows,
        "age_total": {
            "applied": total_applied,
            "matched": total_matched,
            "percentage": _percentage(total_matched, total_applied),
        },
    }


def _build_match_view(result: dict[str, Any]) -> dict[str, Any]:
    household_rows: list[dict[str, str | int]] = []
    child_lookup: dict[str, dict[str, Any]] = {}

    for household_id in sorted(result["household_result_dict"], key=_id_sort_key):
        household = result["household_result_dict"][household_id]
        child_ids = list(household.get("child_ids", []))
        selected_combo = household.get("selected_combo", [])
        combo_rank = household.get("combo_rank")

        household_rows.append(
            {
                "household_id": household["household_id"],
                "child_ids": _display_value(child_ids),
                "selected_combo": _display_value(selected_combo),
                "combo_rank": _display_value(combo_rank),
            }
        )
        for child_id in child_ids:
            child_lookup[child_id] = {
                "household_id": household["household_id"],
                "combo_rank": combo_rank if isinstance(combo_rank, int) else None,
            }

    assignment_rows: list[dict[str, str | int]] = []
    for child_id in sorted(result["matching_result_dict"], key=_id_sort_key):
        assignment = result["matching_result_dict"][child_id]
        household_meta = child_lookup.get(child_id, {})
        assignment_rows.append(
            {
                "child_id": child_id,
                "household_id": household_meta.get("household_id", "-"),
                "daycare_id": _display_value(assignment),
                "combo_rank": _display_value(household_meta.get("combo_rank")),
            }
        )

    is_success = result.get("meta", {}).get("is_optimal", False)
    return {
        "match_summary": _match_summary(result, len(assignment_rows)),
        "household_rows": household_rows,
        "assignment_rows": assignment_rows,
        "is_success": is_success,
        "status_label": "マッチング完了" if is_success else "マッチング不完全",
        "breakdown": _build_breakdown(result.get("matched_children", {})),
    }


_JST = timezone(timedelta(hours=9))


def _build_matching_excel(result: dict[str, Any], meta: dict[str, str] | None = None) -> str:
    output_columns: list[str] = result.get("output_columns", [])
    output_rows: list[dict[str, Any]] = result.get("output_rows", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "result"
    ws.append(output_columns)
    for row in output_rows:
        ws.append([row.get(col, "") for col in output_columns])
    # Prevent formula injection: strings starting with '=' are detected as
    # formula cells by openpyxl; force them back to string type so Excel
    # renders them as literal text instead of executing them.
    for ws_row in ws.iter_rows():
        for cell in ws_row:
            if cell.data_type == "f":
                cell.data_type = "s"

    meta_ws = wb.create_sheet(title="meta")
    meta_ws.append(["key", "value"])
    meta_items: list[tuple[str, str]] = [
        ("schema_version", str(SCHEMA_VERSION)),
        ("active_profile", (meta or {}).get("active_profile", "")),
        ("executed_at", (meta or {}).get("executed_at", "")),
    ]
    for key, value in meta_items:
        meta_ws.append([key, value])
    for ws_row in meta_ws.iter_rows():
        for cell in ws_row:
            if cell.data_type == "f":
                cell.data_type = "s"

    buffer = io.BytesIO()
    wb.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


class ValidationErrorItem(BaseModel):
    model_config = ConfigDict(
        json_schema_extra={
            "examples": [
                {"message": "子どもファイルに必要な列が見つかりません。", "type": "config", "code": 101}
            ]
        }
    )

    message: str = Field(description="エラー内容の説明。")
    type: Literal["config", "data", "format"] = Field(
        description="エラー種別。config は列名マッピング等の設定、data は入力データ、format はファイル形式や列構造の問題を表します。"
    )
    code: int | None = Field(default=None, description="ChilmAI のエラーコード。未分類の場合は null です。")


class ValidateResponse(BaseModel):
    model_config = ConfigDict(
        json_schema_extra={
            "examples": [
                {
                    "is_valid": True,
                    "errors": [],
                    "warnings": [],
                    "summary": {"children_count": 3, "daycares_count": 2},
                }
            ]
        }
    )

    is_valid: bool = Field(description="入力データがマッチング実行可能な状態かどうか。")
    errors: list[ValidationErrorItem] = Field(description="検証で見つかったエラー一覧。")
    warnings: list[str] = Field(description="マッチング実行は可能だが確認を推奨する注意事項。")
    summary: dict[str, Any] = Field(description="申込者数、保育所数などの検証サマリ。")


class ByAgeBucket(BaseModel):
    applied: int = Field(default=0, description="この年齢区分の申込者数。")
    matched: int = Field(default=0, description="この年齢区分で割り当てられた児童数。")


class MatchedChildren(BaseModel):
    total: int = Field(default=0, description="入所決定した児童数。転園元復帰は含みません。")
    only_child: int = Field(default=0, description="単独児世帯で入所決定した児童数。")
    siblings: int = Field(default=0, description="きょうだい世帯で入所決定した児童数。")
    applied_total: int = Field(default=0, description="マッチング対象の申込者総数。")
    by_age: dict[str, ByAgeBucket] = Field(
        default_factory=dict, description="年齢区分ごとの申込者数と割当数。キーは年齢です。"
    )


class HouseholdResult(BaseModel):
    household_id: str = Field(description="世帯 ID。")
    child_ids: list[str] = Field(description="世帯に含まれる児童 ID の一覧。")
    assigned: list[str | None] = Field(description="児童ごとの割当先保育所 ID。未割当の場合は null です。")
    selected_combo: list[str | None] = Field(
        description="選択された児童ごとの保育所組み合わせ。未割当の場合は null です。"
    )
    combo_rank: int | None = Field(
        default=None,
        description="選択された組み合わせの順位。選択されなかった場合は null です。",
    )


class MatchResponse(BaseModel):
    model_config = ConfigDict(
        json_schema_extra={
            "examples": [
                {
                    "matching_result_dict": {"1": "100", "2": None},
                    "household_result_dict": {
                        "10": {
                            "household_id": "10",
                            "child_ids": ["1"],
                            "assigned": ["100"],
                            "selected_combo": ["100"],
                            "combo_rank": 1,
                        }
                    },
                    "matched_children": {
                        "total": 1,
                        "only_child": 1,
                        "siblings": 0,
                        "applied_total": 2,
                        "by_age": {"1": {"applied": 2, "matched": 1}},
                    },
                    "meta": {"algorithm": "cp_use_transfer", "is_optimal": True},
                    "transfer_back_count": 0,
                }
            ]
        }
    )

    matching_result_dict: dict[str, str | None] = Field(
        description="児童 ID から割当先保育所 ID への対応。未割当の場合は null です。"
    )
    household_result_dict: dict[str, HouseholdResult] = Field(description="世帯単位の選考結果。")
    matched_children: MatchedChildren = Field(description="割当件数のサマリ。")
    meta: dict[str, Any] = Field(description="使用アルゴリズムや最適性などのメタ情報。")
    transfer_back_count: int = Field(default=0, description="転園不成立により転園元へ戻った児童数。")


class ConfigResponse(BaseModel):
    model_config = ConfigDict(
        json_schema_extra={
            "examples": [
                {
                    "children": {
                        "child_id": "申請者番号",
                        "household_id": "世帯番号",
                        "age": "年齢",
                        "score_prefix": "点数N",
                        "preference_prefix": "希望保育園ID_N",
                        "enrolled_daycare_id": "在籍保育所ID",
                        "sibling_pattern": "きょうだいパターン",
                    },
                    "daycares": {
                        "daycare_id": "保育所ID",
                        "daycare_name": "保育所名",
                        "capacity_prefix": "N歳募集人数",
                    },
                    "output": {
                        "result_daycare_id": "入所選考結果保育所ID",
                        "result_daycare_name": "入所選考結果保育所名",
                        "exclude_transfer_back": "false",
                    },
                }
            ]
        }
    )

    children: dict[str, str] = Field(
        description="申込者データの内部項目名から実ファイル列名へのマッピング。"
    )
    daycares: dict[str, str] = Field(
        description="保育所データの内部項目名から実ファイル列名へのマッピング。"
    )
    output: dict[str, str] = Field(
        default_factory=dict, description="結果ファイルの出力列名や出力挙動の設定。"
    )
    combination: dict[str, str] = Field(
        default_factory=dict,
        description="組み合わせデータの列名マッピング。任意きょうだい組み合わせを使う場合のみ設定する。",
    )


def _check_same_origin(request: Request) -> None:
    origin = request.headers.get("origin", "")
    referer = request.headers.get("referer", "")
    host = request.headers.get("host", "")
    source = origin or referer
    if not source or not host:
        return
    parsed = urlparse(source)
    if parsed.scheme not in ("http", "https") or parsed.netloc != host:
        raise HTTPException(status_code=403, detail="Forbidden")


@app.get("/", response_class=HTMLResponse, summary="Web UI トップ画面", include_in_schema=False)
def top(request: Request):
    config_display = _get_config_display()
    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "static_version": _static_asset_version(),
            "config_is_customized": config_display["is_customized"],
            "active_profile_name": config_display["active_profile_name"],
        },
    )


@app.post(
    "/config/reset",
    summary="アクティブプロファイルの設定を初期化",
    include_in_schema=False,
)
def reset_config(request: Request) -> RedirectResponse:
    _check_same_origin(request)
    config_store.reset_active()
    return RedirectResponse(url="/settings", status_code=303)


def _settings_context() -> dict[str, Any]:
    snapshot = config_store.snapshot()
    config = snapshot["active_config"]
    profile_names = snapshot["profile_names"]
    active_profile_name = snapshot["active_name"]
    children_fields = [
        (
            key,
            _CHILDREN_LABELS.get(key, key),
            config["children"].get(key, ""),
            _CHILDREN_SAMPLES.get(key, ""),
        )
        for key in _CHILDREN_LABELS
    ]
    daycares_fields = [
        (
            key,
            _DAYCARES_LABELS.get(key, key),
            config["daycares"].get(key, ""),
            _DAYCARES_SAMPLES.get(key, ""),
        )
        for key in _DAYCARES_LABELS
    ]
    output_fields = [
        (key, _OUTPUT_LABELS.get(key, key), config["output"].get(key, ""), _OUTPUT_SAMPLES.get(key, ""))
        for key in _OUTPUT_LABELS
    ]
    combination_fields = [
        (
            key,
            _COMBINATION_LABELS.get(key, key),
            config.get("combination", {}).get(key, ""),
            _COMBINATION_SAMPLES.get(key, ""),
        )
        for key in _COMBINATION_LABELS
    ]
    return {
        "children_fields": children_fields,
        "daycares_fields": daycares_fields,
        "output_fields": output_fields,
        "combination_fields": combination_fields,
        "exclude_transfer_back": config["output"].get("exclude_transfer_back", "false"),
        "profile_names": profile_names,
        "active_profile_name": active_profile_name,
    }


_FLASH_COOKIE_MESSAGE = "profile_flash_message"
_FLASH_COOKIE_ERROR = "profile_flash_error"


@app.get("/settings", response_class=HTMLResponse, summary="項目名設定画面", include_in_schema=False)
def settings(request: Request):
    context = _settings_context()
    context["static_version"] = _static_asset_version()
    message_raw = request.cookies.get(_FLASH_COOKIE_MESSAGE)
    error_raw = request.cookies.get(_FLASH_COOKIE_ERROR)
    context["profile_message"] = unquote(message_raw) if message_raw else None
    context["profile_error"] = unquote(error_raw) if error_raw else None
    response = templates.TemplateResponse(request, "settings.html", context)
    if message_raw is not None:
        response.delete_cookie(_FLASH_COOKIE_MESSAGE, path="/")
    if error_raw is not None:
        response.delete_cookie(_FLASH_COOKIE_ERROR, path="/")
    return response


@app.get(
    "/config",
    response_model=ConfigResponse,
    tags=["API"],
    summary="現在の列名マッピングを取得",
    description="アクティブプロファイルの children / daycares / output 設定を返します。",
)
def get_config() -> ConfigResponse:
    return ConfigResponse(**config_store.load())


@app.post(
    "/config",
    response_model=ConfigResponse,
    tags=["API"],
    summary="列名マッピングを更新",
    description="アクティブプロファイルの列名マッピングを保存し、保存後の設定を返します。",
)
def set_config(config: ConfigResponse) -> ConfigResponse:
    saved = config_store.save(config.model_dump())
    return ConfigResponse(**saved)


@app.post(
    "/validate",
    response_model=ValidateResponse,
    tags=["API"],
    summary="申込者データと保育所データを検証",
    description="CSV/XLSX の列名、必須値、希望園、きょうだい条件などを確認し、マッチング実行前の検証結果を返します。",
)
async def validate(
    children_file: UploadFile = File(..., description="申込者データの CSV または Excel ファイル。"),
    daycares_file: UploadFile = File(..., description="保育所データの CSV または Excel ファイル。"),
    combination_file: UploadFile | None = File(
        default=None, description="組み合わせデータの CSV または Excel ファイル（任意）。"
    ),
) -> ValidateResponse:
    mapping = config_store.load()
    combo_bytes = await combination_file.read() if combination_file and combination_file.filename else None
    try:
        result = matching_service.validate(
            children_file_bytes=await children_file.read(),
            children_file_format=_infer_file_format_from_filename(children_file.filename),
            daycares_file_bytes=await daycares_file.read(),
            daycares_file_format=_infer_file_format_from_filename(daycares_file.filename),
            mapping=mapping,
            combination_file_bytes=combo_bytes,
            combination_file_format=(
                _infer_file_format_from_filename(combination_file.filename)
                if combination_file and combination_file.filename
                else None
            ),
        )
    except ValueError as e:
        if len(e.args) > 1 and isinstance(e.args[1], dict):
            detail: Any = e.args[1]
        else:
            error_code = getattr(e, "code", None)
            error_message = str(e.args[0]) if e.args else str(e)
            detail = {
                "is_valid": False,
                "errors": [
                    {"message": error_message, "type": _error_type_for_code(error_code), "code": error_code}
                ],
                "warnings": [],
                "summary": {},
            }
        raise HTTPException(status_code=422, detail=detail)
    return ValidateResponse(**result)


@app.post(
    "/match",
    response_model=MatchResponse,
    tags=["API"],
    summary="入所選考マッチングを実行",
    description=(
        "申込者データと保育所データを検証したうえで、CP-SAT によるマッチングを実行します。"
        "`solver_config` には JSON 文字列で `max_time_seconds` などを指定できます。"
    ),
)
async def match(
    children_file: UploadFile = File(..., description="申込者データの CSV または Excel ファイル。"),
    daycares_file: UploadFile = File(..., description="保育所データの CSV または Excel ファイル。"),
    combination_file: UploadFile | None = File(
        default=None, description="組み合わせデータの CSV または Excel ファイル（任意）。"
    ),
    solver_config: str | None = Form(
        default=None, description='ソルバー設定の JSON 文字列。例: {"max_time_seconds": 10}'
    ),
) -> MatchResponse:
    parsed_solver_config: dict[str, Any] = {}
    if solver_config:
        try:
            raw = json.loads(solver_config)
            if isinstance(raw, dict):
                parsed_solver_config = raw
        except json.JSONDecodeError:
            parsed_solver_config = {}
    mapping = config_store.load()
    combo_bytes = await combination_file.read() if combination_file and combination_file.filename else None
    try:
        result = matching_service.match(
            children_file_bytes=await children_file.read(),
            children_file_format=_infer_file_format_from_filename(children_file.filename),
            daycares_file_bytes=await daycares_file.read(),
            daycares_file_format=_infer_file_format_from_filename(daycares_file.filename),
            mapping=mapping,
            solver_config=parsed_solver_config,
            combination_file_bytes=combo_bytes,
            combination_file_format=(
                _infer_file_format_from_filename(combination_file.filename)
                if combination_file and combination_file.filename
                else None
            ),
        )
    except ValueError as e:
        if len(e.args) > 1 and isinstance(e.args[1], dict):
            detail: Any = e.args[1]
        else:
            error_code = getattr(e, "code", None)
            error_message = str(e.args[0]) if e.args else str(e)
            detail = {
                "is_valid": False,
                "errors": [
                    {"message": error_message, "type": _error_type_for_code(error_code), "code": error_code}
                ],
                "warnings": [],
                "summary": {},
            }
        raise HTTPException(status_code=422, detail=detail)
    return MatchResponse(**result)


@app.post(
    "/htmx/validate",
    response_class=HTMLResponse,
    summary="Web UI 用のデータ検証 HTML を返す",
    include_in_schema=False,
)
async def htmx_validate(
    request: Request,
    children_file: UploadFile = File(..., description="申込者データの CSV または Excel ファイル。"),
    daycares_file: UploadFile = File(..., description="保育所データの CSV または Excel ファイル。"),
    combination_file: UploadFile | None = File(
        default=None, description="組み合わせデータの CSV または Excel ファイル（任意）。"
    ),
):
    mapping = config_store.load()
    combo_bytes = await combination_file.read() if combination_file and combination_file.filename else None
    try:
        result = matching_service.validate(
            children_file_bytes=await children_file.read(),
            children_file_format=_infer_file_format_from_filename(children_file.filename),
            daycares_file_bytes=await daycares_file.read(),
            daycares_file_format=_infer_file_format_from_filename(daycares_file.filename),
            mapping=mapping,
            combination_file_bytes=combo_bytes,
            combination_file_format=(
                _infer_file_format_from_filename(combination_file.filename)
                if combination_file and combination_file.filename
                else None
            ),
        )
    except ValueError as e:
        error_code = getattr(e, "code", None)
        error_message = str(e.args[0]) if e.args else str(e)
        details = (
            e.args[1]
            if len(e.args) > 1 and isinstance(e.args[1], dict)
            else {
                "is_valid": False,
                "errors": [
                    {"message": error_message, "type": _error_type_for_code(error_code), "code": error_code}
                ],
                "warnings": [],
                "summary": {},
            }
        )
        return templates.TemplateResponse(
            request,
            "validate_result.html",
            {
                "result": details,
                "view": _build_validate_view(details),
            },
            headers={"X-Result-Mode": "validate-error"},
        )
    response = templates.TemplateResponse(
        request,
        "validate_result.html",
        {
            "result": result,
            "view": _build_validate_view(result),
        },
    )
    response.headers["X-Validation-Status"] = "valid" if result["is_valid"] else "invalid"
    response.headers["X-Result-Mode"] = "validate"
    return response


@app.post(
    "/htmx/match",
    response_class=HTMLResponse,
    summary="Web UI 用のマッチング結果 HTML を返す",
    include_in_schema=False,
)
async def htmx_match(
    request: Request,
    children_file: UploadFile = File(..., description="申込者データの CSV または Excel ファイル。"),
    daycares_file: UploadFile = File(..., description="保育所データの CSV または Excel ファイル。"),
    combination_file: UploadFile | None = File(
        default=None, description="組み合わせデータの CSV または Excel ファイル（任意）。"
    ),
    solver_config: str | None = Form(
        default=None, description='ソルバー設定の JSON 文字列。例: {"max_time_seconds": 10}'
    ),
):
    parsed_solver_config: dict[str, Any] = {}
    if solver_config:
        try:
            raw = json.loads(solver_config)
            if isinstance(raw, dict):
                parsed_solver_config = raw
        except json.JSONDecodeError:
            parsed_solver_config = {}
    mapping = config_store.load()
    combo_bytes = await combination_file.read() if combination_file and combination_file.filename else None
    try:
        result = matching_service.match(
            children_file_bytes=await children_file.read(),
            children_file_format=_infer_file_format_from_filename(children_file.filename),
            daycares_file_bytes=await daycares_file.read(),
            daycares_file_format=_infer_file_format_from_filename(daycares_file.filename),
            mapping=mapping,
            solver_config=parsed_solver_config,
            combination_file_bytes=combo_bytes,
            combination_file_format=(
                _infer_file_format_from_filename(combination_file.filename)
                if combination_file and combination_file.filename
                else None
            ),
        )
        active_profile_name = config_store.get_active_name()
        executed_at = datetime.now(_JST).strftime("%Y-%m-%d %H:%M:%S")
        excel_meta = {
            "active_profile": active_profile_name,
            "executed_at": executed_at,
        }
        return templates.TemplateResponse(
            request,
            "match_result.html",
            {
                "result": result,
                "view": _build_match_view(result),
                "matching_excel_b64": _build_matching_excel(result, excel_meta),
                "active_profile_name": active_profile_name,
                "executed_at": executed_at,
            },
            headers={"X-Result-Mode": "match"},
        )
    except OptimizationFailureError as e:
        return templates.TemplateResponse(
            request,
            "match_error.html",
            {"view": {"error_message": str(e), "error_code": getattr(e, "code", None)}},
            headers={"X-Result-Mode": "solver-error"},
        )
    except ValueError as e:
        error_code = getattr(e, "code", None)
        error_message = str(e.args[0]) if e.args else str(e)
        details = (
            e.args[1]
            if len(e.args) > 1 and isinstance(e.args[1], dict)
            else {
                "is_valid": False,
                "errors": [
                    {"message": error_message, "type": _error_type_for_code(error_code), "code": error_code}
                ],
                "warnings": [],
                "summary": {},
            }
        )
        return templates.TemplateResponse(
            request,
            "validate_result.html",
            {
                "result": details,
                "view": _build_validate_view(details),
            },
            headers={"X-Result-Mode": "validate-error"},
        )


_MATCH_TYPE_LABELS: dict[str, str] = {
    "exact": "完全一致",
    "partial": "部分一致",
    "fuzzy": "類似",
    "template": "パターン一致",
}


def _suggestion_to_view(s, section_label: str, label_map: dict[str, str]) -> dict[str, Any]:
    return {
        "section": s.section,
        "section_label": section_label,
        "internal_key": s.internal_key,
        "internal_label": label_map.get(s.internal_key, s.internal_key),
        "current_value": s.current_value,
        "detected_value": s.detected_value,
        "match_type": s.match_type,
        "match_label": _MATCH_TYPE_LABELS.get(s.match_type, s.match_type),
        "confidence": s.confidence,
        "field_name": f"{s.section}__{s.internal_key}",
    }


@app.post(
    "/htmx/analyze-columns",
    response_class=HTMLResponse,
    summary="Web UI 用の列名自動検出 HTML を返す",
    include_in_schema=False,
)
async def htmx_analyze_columns(
    request: Request,
    children_file: UploadFile | None = File(default=None, description="列名を解析する申込者データ。"),
    daycares_file: UploadFile | None = File(default=None, description="列名を解析する保育所データ。"),
):
    current = config_store.load()
    suggestions: list[dict[str, Any]] = []
    errors: list[str] = []
    analyzed_files: list[str] = []

    async def _analyze(
        upload: UploadFile | None, section: str, label_map: dict[str, str], section_label: str
    ):
        if upload is None or not upload.filename:
            return
        try:
            columns = InputParser.read_columns(
                upload.file, _infer_file_format_from_filename(upload.filename)
            )
        except ValueError as e:
            error_code = getattr(e, "code", None)
            if error_code == ErrorCode.UNSUPPORTED_FILE_FORMAT:
                # 変換後メッセージがファイル名を含むためプレフィックス不要。
                errors.append(_labeled_error(_format_error_message(str(e), error_code), error_code))
            elif error_code == ErrorCode.CSV_ENCODING_ERROR:
                # メッセージにファイル名が無いため、どのファイルが原因か分かるよう補う。
                errors.append(f"「{upload.filename}」：{_labeled_error(str(e), error_code)}")
            else:
                errors.append(f"「{upload.filename}」の解析に失敗しました。")
            return
        except Exception:
            errors.append(f"「{upload.filename}」の解析に失敗しました。")
            return
        analyzed_files.append(upload.filename)
        for s in detect_section(columns, section, current[section]):  # type: ignore[arg-type]
            suggestions.append(_suggestion_to_view(s, section_label, label_map))

    await _analyze(children_file, "children", _CHILDREN_LABELS, "申込者データ")
    await _analyze(daycares_file, "daycares", _DAYCARES_LABELS, "保育所データ")

    return templates.TemplateResponse(
        request,
        "column_suggestion.html",
        {
            "suggestions": suggestions,
            "errors": errors,
            "analyzed_files": analyzed_files,
        },
    )


@app.post(
    "/htmx/config",
    response_class=HTMLResponse,
    summary="Web UI 用に列名設定を保存し HTML を返す",
    include_in_schema=False,
)
async def htmx_config(request: Request):
    form = await request.form()
    current = config_store.load()
    updated = {
        "children": current["children"].copy(),
        "daycares": current["daycares"].copy(),
        "output": current["output"].copy(),
        "combination": current.get("combination", {}).copy(),
    }

    for key, value in form.items():
        if not isinstance(key, str):
            continue
        if not isinstance(value, str):
            continue
        if key.startswith("children__"):
            updated["children"][key.replace("children__", "", 1)] = value
        elif key.startswith("daycares__"):
            updated["daycares"][key.replace("daycares__", "", 1)] = value
        elif key.startswith("output__"):
            updated["output"][key.replace("output__", "", 1)] = value
        elif key.startswith("combination__"):
            updated["combination"][key.replace("combination__", "", 1)] = value

    saved = config_store.save(updated)
    return templates.TemplateResponse(
        request,
        "config_result.html",
        {
            "config": saved,
            "children_labels": _CHILDREN_LABELS,
            "daycares_labels": _DAYCARES_LABELS,
            "active_profile_name": config_store.get_active_name(),
        },
    )


def _profile_redirect(*, message: str | None = None, error: str | None = None) -> RedirectResponse:
    response = RedirectResponse(url="/settings", status_code=303)
    if message:
        response.set_cookie(
            _FLASH_COOKIE_MESSAGE,
            quote(message, safe=""),
            max_age=10,
            httponly=True,
            samesite="lax",
            path="/",
        )
    if error:
        response.set_cookie(
            _FLASH_COOKIE_ERROR,
            quote(error, safe=""),
            max_age=10,
            httponly=True,
            samesite="lax",
            path="/",
        )
    return response


def _run_profile_action(request: Request, action: Callable[[], str]) -> RedirectResponse:
    _check_same_origin(request)
    try:
        message = action()
    except ValueError as e:
        return _profile_redirect(error=str(e))
    return _profile_redirect(message=message)


@app.post(
    "/profiles/activate",
    summary="設定プロファイルを切り替える",
    include_in_schema=False,
)
def profiles_activate(
    request: Request,
    profile_name: str = Form(..., description="切り替え先のプロファイル名。"),
) -> RedirectResponse:
    def do_action() -> str:
        new_active = config_store.set_active(profile_name)
        return f"プロファイル「{new_active}」に切り替えました。"

    return _run_profile_action(request, do_action)


@app.post(
    "/profiles/create",
    summary="設定プロファイルを作成する",
    include_in_schema=False,
)
def profiles_create(
    request: Request,
    profile_name: str = Form(..., description="作成するプロファイル名。"),
    copy_from: str | None = Form(
        default=None, description="コピー元プロファイル名。未指定の場合はデフォルト設定から作成します。"
    ),
) -> RedirectResponse:
    def do_action() -> str:
        source = copy_from.strip() if isinstance(copy_from, str) and copy_from.strip() else None
        created = config_store.create_profile(profile_name, source=source)
        config_store.set_active(created)
        if source:
            return f"プロファイル「{created}」を「{source}」から作成し、切り替えました。"
        return f"プロファイル「{created}」を新規作成し、切り替えました。"

    return _run_profile_action(request, do_action)


@app.post(
    "/profiles/rename",
    summary="設定プロファイル名を変更する",
    include_in_schema=False,
)
def profiles_rename(
    request: Request,
    old_name: str = Form(..., description="変更前のプロファイル名。"),
    new_name: str = Form(..., description="変更後のプロファイル名。"),
) -> RedirectResponse:
    def do_action() -> str:
        renamed = config_store.rename_profile(old_name, new_name)
        return f"プロファイル名を「{renamed}」に変更しました。"

    return _run_profile_action(request, do_action)


@app.post(
    "/profiles/delete",
    summary="設定プロファイルを削除する",
    include_in_schema=False,
)
def profiles_delete(
    request: Request,
    profile_name: str = Form(..., description="削除するプロファイル名。"),
) -> RedirectResponse:
    def do_action() -> str:
        new_active = config_store.delete_profile(profile_name)
        return f"プロファイル「{profile_name}」を削除しました（現在: 「{new_active}」）。"

    return _run_profile_action(request, do_action)
