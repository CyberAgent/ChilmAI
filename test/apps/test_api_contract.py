from __future__ import annotations

import base64
import io
import json
import re
from unittest.mock import patch

import openpyxl

import pytest

fastapi = pytest.importorskip("fastapi")
from fastapi.testclient import TestClient  # noqa: E402

from apps.api.main import _infer_file_format_from_filename, app  # noqa: E402
from chilmai.generic.error_codes import ChilmError, ErrorCode  # noqa: E402


def _children_csv() -> bytes:
    return (
        "child_id,household_id,age,score_1,pref_1,pref_2,enrolled_daycare_id,sibling_pattern\n"
        "1,10,1,100,100,101,,\n"
        "2,11,1,80,100,101,,\n"
        "3,12,1,60,101,100,,\n"
    ).encode("utf-8")


def _daycares_csv() -> bytes:
    return (
        "daycare_id,daycare_name,capacity_age0,capacity_age1,capacity_age2,capacity_age3,capacity_age4,capacity_age5\n"
        "100,A,0,1,0,0,0,0\n"
        "101,B,0,1,0,0,0,0\n"
    ).encode("utf-8")


def _children_jp_csv() -> bytes:
    return ("申請者ID,世帯ID,年齢,スコア1,スコア2,希望保育園ID_1\n" "1,10,1,100,90,100\n").encode("utf-8")


def _daycares_jp_csv() -> bytes:
    return ("園ID,園名,0歳募集人数,1歳募集人数\n" "100,A,0,1\n").encode("utf-8")


@pytest.mark.parametrize(
    ("filename", "expected"),
    [
        ("children.csv", "csv"),
        ("children.CSV", "csv"),
        ("daycares.xlsx", "excel"),
        ("daycares.xls", "excel"),
        ("daycares.xlsm", "excel"),
        ("daycares.xlsb", "excel"),
    ],
)
def test_infer_file_format_from_filename(filename: str, expected: str):
    assert _infer_file_format_from_filename(filename) == expected


def test_infer_file_format_from_filename_rejects_unsupported_extension():
    with pytest.raises(ChilmError) as exc_info:
        _infer_file_format_from_filename("IMG_0841.jpg")
    assert exc_info.value.code == ErrorCode.UNSUPPORTED_FILE_FORMAT


def test_analyze_columns_returns_suggestions():
    client = TestClient(app)
    resp = client.post(
        "/htmx/analyze-columns",
        files={
            "children_file": ("c.csv", io.BytesIO(_children_jp_csv()), "text/csv"),
            "daycares_file": ("d.csv", io.BytesIO(_daycares_jp_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    body = resp.text
    assert 'data-target-name="children__child_id"' in body
    assert 'data-detected-value="申請者ID"' in body
    assert 'data-target-name="daycares__daycare_id"' in body
    assert 'data-detected-value="園ID"' in body
    assert "suggestion-accept-checkbox" in body


def test_analyze_columns_unsupported_format_returns_error():
    client = TestClient(app)
    resp = client.post(
        "/htmx/analyze-columns",
        files={"children_file": ("bad.txt", io.BytesIO(b"foo"), "text/plain")},
    )
    assert resp.status_code == 200
    assert "notice--error" in resp.text
    assert "対応していない形式" in resp.text
    assert "E401" in resp.text  # コード付き表示（validate 画面と整合）


def test_analyze_columns_shiftjis_csv_shows_filename_and_code():
    """Shift-JIS CSV の解析失敗時、ファイル名と E405 コード付きで表示されること。"""
    client = TestClient(app)
    shiftjis = "申請者ID,希望保育園ID_1\n1,100\n".encode("cp932")
    resp = client.post(
        "/htmx/analyze-columns",
        files={"children_file": ("kodomo.csv", io.BytesIO(shiftjis), "text/csv")},
    )
    assert resp.status_code == 200
    assert "notice--error" in resp.text
    assert "kodomo.csv" in resp.text  # どのファイルが原因か分かる
    assert "E405" in resp.text  # コード付き表示（validate 画面と整合）


def test_analyze_columns_no_files_returns_empty_state():
    client = TestClient(app)
    resp = client.post("/htmx/analyze-columns", files={})
    assert resp.status_code == 200
    assert "解析するファイルが選択されていません" in resp.text


def test_openapi_documents_api_group_only():
    client = TestClient(app)
    resp = client.get("/openapi.json")
    assert resp.status_code == 200
    schema = resp.json()

    tag_names = {tag["name"] for tag in schema["tags"]}
    assert tag_names == {"API"}
    api_tag = next(tag for tag in schema["tags"] if tag["name"] == "API")
    assert api_tag["description"] == "評価や連携検証向けの API です。"

    paths = schema["paths"]
    assert paths["/validate"]["post"]["tags"] == ["API"]
    assert paths["/match"]["post"]["tags"] == ["API"]
    assert paths["/config"]["get"]["tags"] == ["API"]
    assert paths["/config"]["post"]["tags"] == ["API"]
    assert "/" not in paths
    assert "/settings" not in paths
    assert "/config/reset" not in paths
    assert "/htmx/validate" not in paths
    assert "/htmx/match" not in paths
    assert "/htmx/analyze-columns" not in paths
    assert "/htmx/config" not in paths
    assert "/profiles/activate" not in paths
    assert "/profiles/create" not in paths
    assert "/profiles/rename" not in paths
    assert "/profiles/delete" not in paths

    match_body = paths["/match"]["post"]["requestBody"]["content"]["multipart/form-data"]["schema"]
    if "$ref" in match_body:
        ref_name = match_body["$ref"].rsplit("/", 1)[-1]
        match_body = schema["components"]["schemas"][ref_name]
    assert "solver_config" in match_body["properties"]
    assert paths["/match"]["post"]["summary"] == "入所選考マッチングを実行"

    components = schema["components"]["schemas"]
    match_response = components["MatchResponse"]
    example = match_response["examples"][0]
    assert example["household_result_dict"]["10"]["assigned"] == ["100"]

    household_schema = components["HouseholdResult"]
    assigned_schema = household_schema["properties"]["assigned"]
    assert assigned_schema["type"] == "array"
    assert match_response["properties"]["household_result_dict"]["additionalProperties"]["$ref"].endswith(
        "/HouseholdResult"
    )


def test_redoc_is_disabled():
    client = TestClient(app)
    resp = client.get("/redoc")
    assert resp.status_code == 404


def test_analyze_columns_matching_current_config_no_suggestions():
    client = TestClient(app)
    matching_csv = ("申請者番号,世帯番号,年齢,点数1,点数2,希望保育園ID_1\n" "1,10,1,100,90,100\n").encode(
        "utf-8"
    )
    resp = client.post(
        "/htmx/analyze-columns",
        files={"children_file": ("c.csv", io.BytesIO(matching_csv), "text/csv")},
    )
    assert resp.status_code == 200
    assert "現在の設定と一致しています" in resp.text


def test_validate_endpoint_success():
    client = TestClient(app)
    resp = client.post(
        "/validate",
        files={
            "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["is_valid"] is True
    assert body["errors"] == []


def test_match_endpoint_cp_sat():
    client = TestClient(app)

    resp = client.post(
        "/match",
        data={"solver_config": json.dumps({"max_time_seconds": 3})},
        files={
            "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["meta"]["algorithm"] == "cp_use_transfer"
    assert "household_result_dict" in body
    assert all(
        isinstance(household["assigned"], list) for household in body["household_result_dict"].values()
    )


def test_match_endpoint_transfer_back_count_in_response():
    """転園元復帰が生じた場合、MatchResponse に transfer_back_count が含まれること。"""
    client = TestClient(app)
    # child 1: enrolled D100, prefers D101 (cap=0) → can't transfer → returns to D100
    # child 2: no enrolled, prefers D101 (cap=0) → unmatched
    children = (
        "child_id,household_id,age,score_1,pref_1,enrolled_daycare_id,sibling_pattern\n"
        "1,10,1,100,101,100,\n"
        "2,20,1,100,101,,\n"
    ).encode("utf-8")
    daycares = (
        "daycare_id,daycare_name,capacity_age0,capacity_age1,capacity_age2,capacity_age3,capacity_age4,capacity_age5\n"
        "100,A,0,0,0,0,0,0\n"
        "101,B,0,0,0,0,0,0\n"
    ).encode("utf-8")
    resp = client.post(
        "/match",
        data={"solver_config": json.dumps({"max_time_seconds": 3})},
        files={
            "children_file": ("children.csv", io.BytesIO(children), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(daycares), "text/csv"),
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["transfer_back_count"] == 1
    assert body["matched_children"]["total"] == 0


def test_match_endpoint_returns_by_age_breakdown():
    """MatchResponse.matched_children に by_age と applied 系フィールドが含まれること。"""
    client = TestClient(app)
    resp = client.post(
        "/match",
        data={"solver_config": json.dumps({"max_time_seconds": 3})},
        files={
            "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    mc = resp.json()["matched_children"]
    assert "by_age" in mc
    assert "applied_total" in mc
    assert mc["by_age"]["1"]["applied"] >= 1


def test_htmx_match_renders_breakdown_tables():
    """マッチング結果画面に年齢別の内訳テーブルが表示されること。"""
    client = TestClient(app)
    resp = client.post(
        "/htmx/match",
        files={
            "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    assert 'class="match-breakdown"' in resp.text
    assert "年齢別" in resp.text
    assert "0歳" in resp.text
    assert "5歳" in resp.text


def test_htmx_results_render_summary_and_download_button():
    client = TestClient(app)

    validate_resp = client.post(
        "/htmx/validate",
        files={
            "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert validate_resp.status_code == 200
    assert 'class="status-chip" data-status="success"' in validate_resp.text
    assert 'class="validate-summary-cards"' in validate_resp.text
    assert "申込者数" in validate_resp.text
    assert 'class="validate-summary-card" data-status="success"' in validate_resp.text
    assert validate_resp.headers["X-Result-Mode"] == "validate"

    match_resp = client.post(
        "/htmx/match",
        files={
            "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert match_resp.status_code == 200
    assert 'id="matching-result-excel"' in match_resp.text
    assert "Excelをダウンロード" in match_resp.text
    assert "matching_result.xlsx" in match_resp.text
    assert 'class="status-chip" data-status="success"' in match_resp.text
    assert "マッチング完了" in match_resp.text
    assert match_resp.headers["X-Result-Mode"] == "match"
    # Excelデータをbase64デコードして読み込む
    b64_match = re.search(r'id="matching-result-excel"[^>]*>([^<]+)<', match_resp.text)
    assert b64_match, "Excelのbase64データがレスポンスに含まれていません"
    excel_bytes = base64.b64decode(b64_match.group(1).strip())
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    # ヘッダー: 元ファイルの列 + 入所選考結果列が末尾に含まれること
    assert headers == [
        "child_id",
        "household_id",
        "age",
        "score_1",
        "pref_1",
        "pref_2",
        "enrolled_daycare_id",
        "sibling_pattern",
        "入所選考結果保育所ID",
        "入所選考結果保育所名",
    ]
    # データ行: 元データ値 + 割当結果が含まれること（日本語列名が文字化けせず読み込めること）
    data_rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    assert any(row == ["1", "10", "1", "100", "100", "101", None, None, "100", "A"] for row in data_rows)


def test_validate_endpoint_error_structure():
    client = TestClient(app)
    invalid_children = "child_id,household_id,age\n1,10,1\n".encode("utf-8")
    resp = client.post(
        "/validate",
        files={
            "children_file": ("children.csv", io.BytesIO(invalid_children), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["is_valid"] is False
    assert isinstance(body["errors"], list)
    assert len(body["errors"]) > 0
    for error in body["errors"]:
        assert "message" in error
        assert "type" in error
        assert error["type"] in ("config", "data")


def test_htmx_validate_renders_categorized_errors():
    """バリデーションエラーが設定エラー／データエラーに分類されて表示される。"""
    client = TestClient(app)
    invalid_children = "child_id,household_id,age\n1,10,1\n".encode("utf-8")
    resp = client.post(
        "/htmx/validate",
        files={
            "children_file": ("children.csv", io.BytesIO(invalid_children), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    assert resp.headers["X-Result-Mode"] == "validate"
    assert resp.headers["X-Validation-Status"] == "invalid"
    assert 'data-status="error"' in resp.text
    assert "設定の問題" in resp.text


def test_match_endpoint_validation_failure_returns_422_with_details():
    """バリデーション失敗時に 422 でエラー詳細が返る。"""
    client = TestClient(app)
    invalid_children = "child_id,household_id,age\n1,10,1\n".encode("utf-8")
    resp = client.post(
        "/match",
        files={
            "children_file": ("children.csv", io.BytesIO(invalid_children), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert isinstance(detail, dict)
    assert detail.get("is_valid") is False
    assert isinstance(detail.get("errors"), list)
    assert len(detail["errors"]) > 0


def test_match_endpoint_value_error_returns_422():
    """`solver_config` に不正な値を渡して ValueError が発生しても 500 にならず 422 を返す。"""
    client = TestClient(app)
    resp = client.post(
        "/match",
        data={"solver_config": json.dumps({"max_time_seconds": "abc"})},
        files={
            "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 422


def test_htmx_match_non_validation_error_renders_gracefully():
    """`solver_config` に不正な値を渡して ValidationError 以外の ValueError が発生しても 500 にならない。"""
    client = TestClient(app)
    resp = client.post(
        "/htmx/match",
        data={"solver_config": json.dumps({"max_time_seconds": "abc"})},
        files={
            "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    assert resp.headers["X-Result-Mode"] == "validate-error"
    assert 'data-status="error"' in resp.text


def test_htmx_match_solver_error_renders_match_error_template():
    """`OptimizationFailureError` が発生したとき `match_error.html` が返り `solver-error` ヘッダが付く。"""
    from chilmai.algorithm.cp_use_transfer.CP_algo import OptimizationFailureError

    client = TestClient(app)
    with patch(
        "apps.api.main.matching_service.match", side_effect=OptimizationFailureError("テストエラー")
    ):
        resp = client.post(
            "/htmx/match",
            files={
                "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
                "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
            },
        )
    assert resp.status_code == 200
    assert resp.headers["X-Result-Mode"] == "solver-error"
    assert 'data-status="error"' in resp.text
    assert "テストエラー" in resp.text


def test_match_service_raises_on_output_column_conflict():
    """出力列名が申込者ファイルの既存列と重複する場合に ValueError が発生する。"""
    from chilmai.generic.service import MatchingService

    service = MatchingService()
    with pytest.raises(ValueError, match="重複"):
        service.match(
            children_file_bytes=_children_csv(),
            children_file_format="csv",
            daycares_file_bytes=_daycares_csv(),
            daycares_file_format="csv",
            mapping={
                "children": {},
                "daycares": {},
                "output": {
                    "result_daycare_id": "child_id",
                    "result_daycare_name": "入所選考結果保育所名",
                },
            },
        )


def test_match_service_raises_on_output_columns_duplicate():
    """出力列名 2 つが同じ値の場合に ValueError が発生する。"""
    from chilmai.generic.service import MatchingService

    service = MatchingService()
    with pytest.raises(ValueError, match="重複"):
        service.match(
            children_file_bytes=_children_csv(),
            children_file_format="csv",
            daycares_file_bytes=_daycares_csv(),
            daycares_file_format="csv",
            mapping={
                "children": {},
                "daycares": {},
                "output": {
                    "result_daycare_id": "入所選考結果",
                    "result_daycare_name": "入所選考結果",
                },
            },
        )


def test_validate_endpoint_unsupported_file_format_returns_422():
    """サポート外ファイル形式（JPG等）をアップロードしても 500 にならず 422 を返す。"""
    client = TestClient(app)
    fake_image = b"\xff\xd8\xff\xe0fake jpeg bytes"
    resp = client.post(
        "/validate",
        files={
            "children_file": ("IMG_0841.jpg", io.BytesIO(fake_image), "image/jpeg"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert isinstance(detail, dict)
    assert detail["is_valid"] is False
    assert any("IMG_0841.jpg" in e["message"] for e in detail["errors"])
    assert detail["errors"][0]["code"] == 401


def test_htmx_validate_unsupported_file_format_renders_error():
    """サポート外ファイル形式（JPG等）をアップロードしても 500 にならずエラー画面を返す。"""
    client = TestClient(app)
    fake_image = b"\xff\xd8\xff\xe0fake jpeg bytes"
    resp = client.post(
        "/htmx/validate",
        files={
            "children_file": ("IMG_0841.jpg", io.BytesIO(fake_image), "image/jpeg"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    assert resp.headers["X-Result-Mode"] == "validate-error"
    assert 'data-status="error"' in resp.text


def test_reset_config_resets_active_profile_and_redirects():
    """POST /config/reset がアクティブプロファイルをデフォルト値に戻し /settings へ 303 リダイレクトする。"""
    client = TestClient(app, follow_redirects=False)
    with patch("apps.api.main.config_store.reset_active") as mock_reset:
        resp = client.post("/config/reset")
        mock_reset.assert_called_once_with()
    assert resp.status_code == 303
    assert resp.headers["location"] == "/settings"


def test_reset_config_rejects_cross_origin():
    """外部 Origin からの POST /config/reset は 403 を返す。"""
    client = TestClient(app)
    resp = client.post(
        "/config/reset",
        headers={"origin": "https://evil.example.com", "host": "localhost:8501"},
    )
    assert resp.status_code == 403


def test_config_roundtrip():
    client = TestClient(app)

    current = client.get("/config")
    assert current.status_code == 200
    body = current.json()
    assert "output" in body
    assert "result_daycare_id" in body["output"]
    assert "result_daycare_name" in body["output"]

    updated = {
        "children": {**body["children"], "child_id": "児童ID"},
        "daycares": body["daycares"],
        "output": {**body["output"], "result_daycare_id": "決定保育所ID"},
    }
    save = client.post("/config", json=updated)
    assert save.status_code == 200
    assert save.json()["children"]["child_id"] == "児童ID"
    assert save.json()["output"]["result_daycare_id"] == "決定保育所ID"

    restore = client.post("/config", json=body)
    assert restore.status_code == 200


def test_htmx_match_solver_error_shows_error_code():
    """OptimizationFailureError に code が付いている場合、HTMX レスポンスに [E5xx] が出力される。"""
    from chilmai.algorithm.cp_use_transfer.CP_algo import OptimizationFailureError
    from chilmai.generic.error_codes import ErrorCode

    client = TestClient(app)
    err = OptimizationFailureError("マッチングが実行不可能です")
    err.code = ErrorCode.SOLVER_INFEASIBLE
    with patch("apps.api.main.matching_service.match", side_effect=err):
        resp = client.post(
            "/htmx/match",
            files={
                "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
                "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
            },
        )
    assert resp.status_code == 200
    assert resp.headers["X-Result-Mode"] == "solver-error"
    assert "[E501]" in resp.text


def test_htmx_validate_error_shows_code_prefix():
    """バリデーションエラーの UI 表示に [E1xx] / [E2xx] プレフィックスが含まれること。"""
    client = TestClient(app)
    invalid_children = "child_id,household_id,age\n1,10,1\n".encode("utf-8")
    resp = client.post(
        "/htmx/validate",
        files={
            "children_file": ("children.csv", io.BytesIO(invalid_children), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    assert "[E1" in resp.text or "[E2" in resp.text


def test_match_endpoint_config_error_returns_config_type():
    """E106（出力列名重複）は type: 'config' で返ること。"""
    from chilmai.generic.error_codes import ChilmError, ErrorCode

    client = TestClient(app)
    err = ChilmError("出力列名の設定が重複しています", code=ErrorCode.DUPLICATE_OUTPUT_COL_NAMES)
    with patch("apps.api.main.matching_service.match", side_effect=err):
        resp = client.post(
            "/match",
            files={
                "children_file": ("children.csv", io.BytesIO(_children_csv()), "text/csv"),
                "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
            },
        )
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert detail["errors"][0]["type"] == "config"
    assert detail["errors"][0]["code"] == ErrorCode.DUPLICATE_OUTPUT_COL_NAMES


def test_validate_endpoint_parser_error_returns_format_type():
    """E403（score 列競合）は type: 'format' で返ること。"""
    client = TestClient(app)
    children_with_score_conflict = (
        "child_id,household_id,age,score,score_1,pref_1,enrolled_daycare_id,sibling_pattern\n"
        "1,10,1,100,100,100,,\n"
    ).encode("utf-8")
    resp = client.post(
        "/validate",
        files={
            "children_file": ("children.csv", io.BytesIO(children_with_score_conflict), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert detail["errors"][0]["type"] == "format"
    assert detail["errors"][0]["code"] == 403


def test_validate_endpoint_error_includes_code_field():
    """/validate エラーレスポンスの errors 各要素に整数の code フィールドが含まれること。"""
    client = TestClient(app)
    invalid_children = "child_id,household_id,age\n1,10,1\n".encode("utf-8")
    resp = client.post(
        "/validate",
        files={
            "children_file": ("children.csv", io.BytesIO(invalid_children), "text/csv"),
            "daycares_file": ("daycares.csv", io.BytesIO(_daycares_csv()), "text/csv"),
        },
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["is_valid"] is False
    for error in body["errors"]:
        assert "code" in error, f"code フィールドがない: {error}"
        assert isinstance(error["code"], int), f"code が int でない: {error['code']!r}"


def test_build_matching_excel_no_formula_injection():
    """入力値に '=' で始まる文字列が含まれても数式セルにならないこと。"""
    from apps.api.main import _build_matching_excel

    result = {
        "output_columns": ["col_a", "col_b"],
        "output_rows": [
            {"col_a": '=HYPERLINK("http://evil.example.com","click")', "col_b": "safe"},
            {"col_a": "+cmd", "col_b": "normal"},
        ],
    }
    b64 = _build_matching_excel(result)
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))
    ws = wb["result"]
    for row in ws.iter_rows():
        for cell in row:
            assert cell.data_type != "f", f"数式セルが検出されました: {cell.coordinate} = {cell.value!r}"


def test_build_matching_excel_meta_sheet_records_profile_and_timestamp():
    """meta シートに active_profile と executed_at が記録される。"""
    from apps.api.main import _build_matching_excel

    result = {"output_columns": ["a"], "output_rows": [{"a": "x"}]}
    meta = {"active_profile": "A市", "executed_at": "2026-05-13 12:00:00"}
    b64 = _build_matching_excel(result, meta)
    wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(b64)))

    assert "meta" in wb.sheetnames
    meta_ws = wb["meta"]
    rows = {row[0].value: row[1].value for row in meta_ws.iter_rows(min_row=2)}
    assert rows.get("active_profile") == "A市"
    assert rows.get("executed_at") == "2026-05-13 12:00:00"
    assert rows.get("schema_version") == "2"


@pytest.fixture
def profiles_client(tmp_path, monkeypatch):
    """data/config.json をテスト専用パスに差し替えた TestClient を返す（テスト後に原状復帰）。"""
    from apps.api import main as api_main
    from chilmai.generic.config import ConfigStore

    store = ConfigStore(tmp_path / "config.json")
    monkeypatch.setattr(api_main, "config_store", store)
    client = TestClient(api_main.app, follow_redirects=False)
    return client, store


def test_profile_create_activate_persists(profiles_client):
    client, store = profiles_client
    resp = client.post("/profiles/create", data={"profile_name": "2026年度"})
    assert resp.status_code == 303
    assert "/settings" in resp.headers["location"]

    assert store.get_active_name() == "2026年度"
    assert "2026年度" in store.list_profiles()


def test_profile_delete_default_returns_error_redirect(profiles_client):
    client, _ = profiles_client
    resp = client.post("/profiles/delete", data={"profile_name": "default"})
    assert resp.status_code == 303
    # URL にエラー文を載せず、ワンショット Cookie で運ぶ。
    assert resp.headers["location"] == "/settings"
    assert "profile_flash_error" in resp.cookies


def test_profile_rename(profiles_client):
    client, store = profiles_client
    client.post("/profiles/create", data={"profile_name": "旧名"})
    resp = client.post("/profiles/rename", data={"old_name": "旧名", "new_name": "新名"})
    assert resp.status_code == 303

    assert "新名" in store.list_profiles()
    assert "旧名" not in store.list_profiles()


def test_profile_delete_active_falls_back_to_default(profiles_client):
    client, store = profiles_client
    client.post("/profiles/create", data={"profile_name": "A市"})
    resp = client.post("/profiles/delete", data={"profile_name": "A市"})
    assert resp.status_code == 303

    assert store.get_active_name() == "default"


def test_profile_endpoints_reject_cross_origin(profiles_client):
    client, _ = profiles_client
    resp = client.post(
        "/profiles/activate",
        data={"profile_name": "default"},
        headers={"origin": "https://evil.example.com", "host": "localhost:8501"},
    )
    assert resp.status_code == 403


def test_profile_endpoints_reject_subdomain_spoof(profiles_client):
    """startswith ベースの旧実装で通り抜けていた host サブドメイン詐称を弾く。"""
    client, _ = profiles_client
    resp = client.post(
        "/profiles/activate",
        data={"profile_name": "default"},
        headers={"origin": "http://localhost.evil.example", "host": "localhost"},
    )
    assert resp.status_code == 403


def test_settings_page_shows_active_profile(tmp_path, monkeypatch):
    from apps.api import main as api_main
    from chilmai.generic.config import ConfigStore

    store = ConfigStore(tmp_path / "config.json")
    monkeypatch.setattr(api_main, "config_store", store)
    store.create_profile("A市")
    store.set_active("A市")

    client = TestClient(api_main.app)
    resp = client.get("/settings")
    assert resp.status_code == 200
    assert "A市" in resp.text
    assert "プロファイル" in resp.text


def test_top_page_shows_active_profile(tmp_path, monkeypatch):
    from apps.api import main as api_main
    from chilmai.generic.config import ConfigStore

    store = ConfigStore(tmp_path / "config.json")
    monkeypatch.setattr(api_main, "config_store", store)
    store.create_profile("Test")
    store.set_active("Test")

    client = TestClient(api_main.app)
    resp = client.get("/")
    assert resp.status_code == 200
    import re

    m = re.search(r'class="config-link__profile"[^>]*>\s*([^<\s]+)\s*<', resp.text)
    assert m is not None, "config-link__profile span not found"
    assert m.group(1) == "Test"


def test_default_profile_renders_as_katakana_on_top_page(tmp_path, monkeypatch):
    """default プロファイルがアクティブな時、ヘッダーのプロファイルチップは「デフォルト」表示。"""
    from apps.api import main as api_main
    from chilmai.generic.config import ConfigStore

    store = ConfigStore(tmp_path / "config.json")
    monkeypatch.setattr(api_main, "config_store", store)

    client = TestClient(api_main.app)
    resp = client.get("/")
    assert resp.status_code == 200
    import re

    m = re.search(r'class="config-link__profile"[^>]*>\s*([^<\s]+)\s*<', resp.text)
    assert m is not None, "config-link__profile span not found"
    assert m.group(1) == "デフォルト"


def test_default_profile_renders_as_katakana_on_settings_page(tmp_path, monkeypatch):
    """設定画面のプロファイル一覧で default は「デフォルト」と表示され、アクティブ行は「編集中」バッジが付く。"""
    from apps.api import main as api_main
    from chilmai.generic.config import ConfigStore

    store = ConfigStore(tmp_path / "config.json")
    monkeypatch.setattr(api_main, "config_store", store)

    client = TestClient(api_main.app)
    resp = client.get("/settings")
    assert resp.status_code == 200
    # 内部キー "default" は activate フォームの hidden input に保持される。
    assert 'name="profile_name" value="default"' in resp.text
    # 表示は「デフォルト」、アクティブ行は「編集中」バッジを伴う。
    assert "デフォルト" in resp.text
    assert "編集中" in resp.text


def test_top_page_is_not_customized_when_only_extra_default_profile(tmp_path, monkeypatch):
    """プロファイルが複数あっても、アクティブが default かつ値もデフォルトなら、未カスタマイズ案内が出る。"""
    from apps.api import main as api_main
    from chilmai.generic.config import ConfigStore

    store = ConfigStore(tmp_path / "config.json")
    monkeypatch.setattr(api_main, "config_store", store)
    store.create_profile("B市")  # 追加するだけ。default はそのままアクティブ＆未編集。

    client = TestClient(api_main.app)
    resp = client.get("/")
    assert resp.status_code == 200
    # 未カスタマイズ時にだけ表示される案内文。
    assert "項目名は" in resp.text and "デフォルト設定" in resp.text
